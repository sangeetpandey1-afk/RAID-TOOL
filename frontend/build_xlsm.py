#!/usr/bin/env python3
"""
Build the starter Excel workbook for the Raid Management System UI.

This script creates ``frontend/RaidSystem.xlsx`` with all required sheets,
headers, named ranges, and instructions in column G/H so the officer
knows where to attach each macro after opening the file.

The user then:
1. Opens it in Excel.
2. Saves As → Excel Macro-Enabled Workbook (.xlsm).
3. Imports each .bas / .cls from ``frontend/vba/`` via Alt+F11 → File → Import.
4. Adds Form Control buttons next to the marker cells and links them to
   the macros listed in the same row.
"""
from __future__ import annotations
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "frontend" / "RaidSystem.xlsx"

# ---------------------------------------------------------------- styles
HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT  = Font(bold=True, size=14, color="1F4E79")
NOTE_FONT   = Font(italic=True, color="606060", size=9)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HINDI_FONT = Font(name="Mangal", size=11)


def style_header_row(ws, row: int, cols: int) -> None:
    for col in range(1, cols + 1):
        c = ws.cell(row=row, column=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)


def autosize(ws, min_w: int = 10, max_w: int = 40) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        m = min_w
        for cell in col:
            if cell.value is not None:
                m = max(m, min(max_w, len(str(cell.value)) + 2))
        ws.column_dimensions[letter].width = m


# ============================================================ sheets
def build_dashboard(wb: Workbook):
    ws = wb.create_sheet("Dashboard", 0)
    ws["A1"] = "विद्युत चोरी रेड प्रबंधन सिस्टम"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Raid Management System — Dashboard"
    ws["A2"].font = NOTE_FONT
    ws.merge_cells("A1:F1")
    ws.merge_cells("A2:F2")

    ws["A4"] = "Backend"
    ws["B4"] = "Status"
    ws["A5"] = "Health"
    ws["B5"] = "(refresh to populate)"

    ws["A7"] = "Total Cases";          ws["B7"] = 0
    ws["A8"] = "Total Assessment ₹";    ws["B8"] = 0
    ws["A9"] = "Today Payments (count)";ws["B9"] = 0
    ws["A10"] = "Today Payment Amt";    ws["B10"] = 0

    # Macro-button placeholder cells
    ws["G4"] = "Action"
    ws["H4"] = "Macro to assign"
    style_header_row(ws, 4, 8)
    actions = [
        ("Refresh Health",      "CheckHealth"),
        ("Import Master Data",  "ImportAllMaster"),
        ("Backup Now",          "RunBackup"),
        ("Export Cases (xlsx)", "ExportCases"),
        ("Export Payments",     "ExportPayments"),
        ("Dashboard PDF",       "ExportDashboardPDF"),
    ]
    for i, (label, macro) in enumerate(actions, start=5):
        ws.cell(row=i, column=7, value=label)
        ws.cell(row=i, column=8, value=macro).font = Font(name="Consolas", size=10)

    ws["A14"] = "Import Report (last run)"
    ws["A14"].font = TITLE_FONT
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["G"].width = 24
    ws.column_dimensions["H"].width = 28


def build_inputs(wb: Workbook):
    ws = wb.create_sheet("Inputs")
    ws["A1"] = "Case Inputs"; ws["A1"].font = TITLE_FONT
    fields = [
        ("Account Number", "nrAccount",       ""),
        ("Name",           "nrName",          ""),
        ("Father / Husband","nrFather",       ""),
        ("Village",        "nrVillage",       ""),
        ("Post Office",    "nrPost",          ""),
        ("Pin Code",       "nrPin",           ""),
        ("Mobile",         "nrMobile",        ""),
        ("Section",        "nrSection",       "135"),
        ("Inspection Date","nrInspectionDate", ""),
        ("Category",       "nrCategory",      "LMV-1"),
        ("Connected Load (kW)", "nrLoadKW",   ""),
        ("J.E. Name",      "nrJE",            ""),
        ("Sub-Substation", "nrSubStation",    ""),
        ("Checking Type",  "nrCheckingType",  "Vigilance"),
    ]
    for i, (label, name, default) in enumerate(fields, start=3):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        c = ws.cell(row=i, column=2, value=default)
        c.border = BORDER
        if "Hindi" in label or label in ("Name", "Father / Husband", "Village"):
            c.font = HINDI_FONT
        # named range
        wb.defined_names[name] = DefinedName(
            name=name, attr_text=f"Inputs!${get_column_letter(2)}${i}")

    ws["G3"] = "Action";  ws["H3"] = "Macro"
    style_header_row(ws, 3, 8)
    inp_actions = [
        ("Live Calculate",       "LiveCalc"),
        ("Save Case",            "SaveCase"),
        ("Offense Check",        "OffenseCheck"),
        ("Generate All Documents","GenerateAllDocuments"),
        ("Generate One Document", "GenerateOneDocument"),
    ]
    for i, (label, macro) in enumerate(inp_actions, start=4):
        ws.cell(row=i, column=7, value=label)
        ws.cell(row=i, column=8, value=macro).font = Font(name="Consolas", size=10)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["G"].width = 26
    ws.column_dimensions["H"].width = 28


def build_devices(wb: Workbook):
    ws = wb.create_sheet("Devices")
    ws["A1"] = "Devices (LFHD)"; ws["A1"].font = TITLE_FONT
    headers = ["Device", "Load (W)", "Factor", "Hours/day", "Days", "Units (auto)"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=2, column=col, value=h)
    style_header_row(ws, 2, len(headers))

    samples = [
        ("Bulb / LED", 9,    1, 6,  365),
        ("Ceiling Fan", 75,  1, 12, 365),
        ("AC 1.5 Ton", 1800, 1, 4,  120),
    ]
    for i, row in enumerate(samples, start=3):
        for j, v in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=v)

    ws["H2"] = "Tip:";  ws["H2"].font = Font(bold=True)
    ws["H3"] = "Add up to 30 devices below row 2."
    ws["H3"].font = NOTE_FONT
    ws.column_dimensions["A"].width = 28
    for col in ("B", "C", "D", "E", "F"):
        ws.column_dimensions[col].width = 12
    ws.column_dimensions["H"].width = 40


def build_calc(wb: Workbook):
    ws = wb.create_sheet("Calc")
    ws["A1"] = "Assessment Breakdown (read-only)"; ws["A1"].font = TITLE_FONT
    ws["A2"] = "Run 'Live Calculate' on Inputs sheet to populate."
    ws["A2"].font = NOTE_FONT
    ws["A4"] = "Energy Charges Slabs"; ws["A4"].font = Font(bold=True)
    ws["A14"] = "Summary"; ws["A14"].font = Font(bold=True)
    ws["A24"] = "Grand Total mirror (named: nrCalcTotal)"
    ws["A24"].font = NOTE_FONT
    ws["B25"] = 0
    wb.defined_names["nrCalcTotal"] = DefinedName(
        name="nrCalcTotal", attr_text="Calc!$B$25")
    autosize(ws)


def build_search(wb: Workbook):
    ws = wb.create_sheet("Search")
    ws["A1"] = "Account"; ws["A2"] = "Name"; ws["A3"] = "Village"
    for r in range(1, 4):
        ws.cell(row=r, column=1).font = Font(bold=True)
        ws.cell(row=r, column=2).border = BORDER

    ws["G1"] = "Action";  ws["H1"] = "Macro"
    style_header_row(ws, 1, 8)
    ws["G2"] = "Search";          ws["H2"] = "SearchConsumer"
    ws["G3"] = "Offense (active)";ws["H3"] = "OffenseCheckOnSelected"

    headers = ["Account", "Name", "Father", "Village", "Mobile", "Category"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=5, column=col, value=h)
    style_header_row(ws, 5, len(headers))

    for col in ("A", "B", "C", "D"):
        ws.column_dimensions[col].width = 18
    ws.column_dimensions["G"].width = 22
    ws.column_dimensions["H"].width = 28


def build_cases(wb: Workbook):
    ws = wb.create_sheet("Cases")
    ws["A1"] = "Current case_id"; ws["A1"].font = Font(bold=True)
    ws["B1"] = ""
    wb.defined_names["nrCurrentCaseId"] = DefinedName(
        name="nrCurrentCaseId", attr_text="Cases!$B$1")

    ws["G1"] = "Action";  ws["H1"] = "Macro"
    style_header_row(ws, 1, 8)
    ws["G2"] = "Refresh List"; ws["H2"] = "RefreshCases"

    headers = ["Case ID", "Account", "Name", "Section", "Inspection",
               "Assessment", "Compounding", "Status"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=3, column=col, value=h)
    style_header_row(ws, 3, len(headers))
    for col in ("A", "B", "C"):
        ws.column_dimensions[col].width = 18


def build_payments(wb: Workbook):
    ws = wb.create_sheet("Payments")
    ws["A1"] = "Payments — refresh to load for current case"
    ws["A1"].font = TITLE_FONT

    ws["G1"] = "Action";  ws["H1"] = "Macro"
    style_header_row(ws, 1, 8)
    ws["G2"] = "Record Payment"; ws["H2"] = "RecordPayment"
    ws["G3"] = "Refresh";        ws["H3"] = "RefreshPayments"

    headers = ["Date", "Amount", "Type", "Component", "Receipt",
               "Method", "Remarks"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=3, column=col, value=h)
    style_header_row(ws, 3, len(headers))
    for col in ("A", "B", "C", "D"):
        ws.column_dimensions[col].width = 16


def build_notices(wb: Workbook):
    ws = wb.create_sheet("Notices")
    ws["A1"] = "Notices Timeline"; ws["A1"].font = TITLE_FONT

    ws["G1"] = "Action";  ws["H1"] = "Macro"
    style_header_row(ws, 1, 8)
    ws["G2"] = "Add Provisional";  ws["H2"] = "AddProvisionalNotice"
    ws["G3"] = "Add Section 3";    ws["H3"] = "AddSection3Notice"
    ws["G4"] = "Add Section 5";    ws["H4"] = "AddSection5Notice"
    ws["G5"] = "Refresh";          ws["H5"] = "RefreshNotices"

    headers = ["Type", "Number", "Date", "Due", "Status",
               "Dispatch", "Created"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=3, column=col, value=h)
    style_header_row(ws, 3, len(headers))
    for col in ("A", "B", "C", "D", "E"):
        ws.column_dimensions[col].width = 16


def build_settings(wb: Workbook):
    ws = wb.create_sheet("Settings")
    ws["A1"] = "Settings"; ws["A1"].font = TITLE_FONT

    rows = [
        ("API Base URL",   "nrApiBase",  "http://127.0.0.1:5000"),
        ("Default Category","nrDefCat",  "LMV-1"),
        ("Multiplier 1st", "nrMult1",    2),
        ("Multiplier Repeat","nrMultR",  6),
    ]
    for i, (label, name, default) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        c = ws.cell(row=i, column=2, value=default)
        c.border = BORDER
        wb.defined_names[name] = DefinedName(
            name=name, attr_text=f"Settings!${get_column_letter(2)}${i}")

    ws["A8"] = "Hindi font hint:"
    ws["B8"] = "Mangal / Nirmala UI / Krutidev 010"
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 36


# ============================================================ main
def build() -> Path:
    wb = Workbook()
    # Remove default sheet
    if wb.active.title == "Sheet":
        wb.remove(wb.active)

    build_dashboard(wb)
    build_inputs(wb)
    build_devices(wb)
    build_calc(wb)
    build_search(wb)
    build_cases(wb)
    build_payments(wb)
    build_notices(wb)
    build_settings(wb)

    # Reorder so Dashboard is first
    wb._sheets.sort(key=lambda s: [
        "Dashboard", "Inputs", "Devices", "Calc", "Search",
        "Cases", "Payments", "Notices", "Settings",
    ].index(s.title))

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"[OK] Wrote: {OUT.relative_to(ROOT)}")
    print("Next: open in Excel, Save As .xlsm, then import all .bas/.cls")
    return OUT


if __name__ == "__main__":
    build()
