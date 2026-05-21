"""
Reports & exports.

Excel via ``openpyxl`` and PDF via ``reportlab``. All outputs land in
``backup/reports/`` so they are excluded from git but easily downloadable
through the ``/api/reports/download/<name>`` endpoint.
"""
from __future__ import annotations
import json
import logging
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .. import config
from ..database import fetch_all
from ..utils import safe_float

log = logging.getLogger(__name__)

REPORTS_DIR: Path = config.BACKUP_DIR / "reports"
REPORTS_DIR.mkdir(parents=True, exist_ok=True)


# ----------------------------------------------------------- excel helpers
_HEADER_FILL = PatternFill("solid", fgColor="305496")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_BODY_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _autosize(ws) -> None:
    for col in ws.columns:
        max_len = 0
        letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:  # noqa: BLE001
                pass
        ws.column_dimensions[letter].width = min(max_len + 2, 50)


def _write_header(ws, headers: list[str]) -> None:
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = _HEADER_ALIGN
    ws.freeze_panes = "A2"


def _stamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


# ===================================================================
# Cases export (xlsx)
# ===================================================================
def cases_xlsx(*, status: str | None = None,
               from_date: str | None = None,
               to_date: str | None = None) -> Path:
    """Export raid_cases joined with consumer name/village to a .xlsx."""
    where = ["1=1"]
    args: list[Any] = []
    if status:
        where.append("rc.case_status = ?")
        args.append(status)
    if from_date:
        where.append("rc.inspection_date >= ?")
        args.append(from_date)
    if to_date:
        where.append("rc.inspection_date <= ?")
        args.append(to_date)

    rows = fetch_all(
        f"""
        SELECT rc.case_id, rc.account_number, c.name, c.father_name,
               c.village, c.div_code, rc.section, rc.inspection_date,
               rc.connected_load_kw, rc.total_assessment,
               rc.compounding_amount, rc.case_status, rc.je_name,
               rc.checking_type, rc.sub_substation, rc.created_at
          FROM raid_cases rc
          LEFT JOIN consumers c ON c.id = rc.consumer_id
         WHERE {' AND '.join(where)}
         ORDER BY rc.created_at DESC
        """,
        args,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Cases"
    headers = ["Case ID", "Account", "Name", "Father", "Village",
               "Div", "Sec", "Inspection", "Load (kW)",
               "Assessment", "Compounding", "Status", "J.E.",
               "Checking Type", "Sub-Substation", "Created"]
    _write_header(ws, headers)
    for r in rows:
        ws.append([
            r["case_id"], r["account_number"], r["name"], r["father_name"],
            r["village"], r["div_code"], r["section"], r["inspection_date"],
            r["connected_load_kw"], r["total_assessment"],
            r["compounding_amount"], r["case_status"], r["je_name"],
            r["checking_type"], r["sub_substation"], r["created_at"],
        ])
    _autosize(ws)

    out = REPORTS_DIR / f"cases_{_stamp()}.xlsx"
    wb.save(out)
    log.info("cases_xlsx: %d rows → %s", len(rows), out.name)
    return out


# ===================================================================
# Payments export (xlsx)
# ===================================================================
def payments_xlsx(*, from_date: str | None = None,
                  to_date: str | None = None) -> Path:
    where = ["1=1"]
    args: list[Any] = []
    if from_date:
        where.append("p.payment_date >= ?")
        args.append(from_date)
    if to_date:
        where.append("p.payment_date <= ?")
        args.append(to_date)

    rows = fetch_all(
        f"""
        SELECT p.id, p.case_id, p.payment_date, p.amount,
               p.payment_type, p.component, p.receipt_number,
               p.payment_method, p.remarks,
               c.name AS consumer_name, c.village,
               rc.account_number
          FROM payments p
          LEFT JOIN raid_cases rc ON rc.case_id = p.case_id
          LEFT JOIN consumers c ON c.id = rc.consumer_id
         WHERE {' AND '.join(where)}
         ORDER BY p.payment_date DESC, p.id DESC
        """,
        args,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Payments"
    headers = ["Pay ID", "Case ID", "Date", "Amount", "Type", "Component",
               "Receipt", "Method", "Remarks", "Consumer", "Village",
               "Account"]
    _write_header(ws, headers)
    total = 0.0
    for r in rows:
        ws.append([
            r["id"], r["case_id"], r["payment_date"], r["amount"],
            r["payment_type"], r["component"], r["receipt_number"],
            r["payment_method"], r["remarks"], r["consumer_name"],
            r["village"], r["account_number"],
        ])
        total += safe_float(r["amount"])

    # Total row
    last = ws.max_row + 2
    ws.cell(row=last, column=3, value="TOTAL").font = Font(bold=True)
    c = ws.cell(row=last, column=4, value=total)
    c.font = Font(bold=True)
    _autosize(ws)

    out = REPORTS_DIR / f"payments_{_stamp()}.xlsx"
    wb.save(out)
    log.info("payments_xlsx: %d rows, total %.2f → %s",
             len(rows), total, out.name)
    return out


# ===================================================================
# Notice timeline (xlsx)
# ===================================================================
def notices_xlsx() -> Path:
    rows = fetch_all(
        """
        SELECT n.id, n.case_id, n.notice_type, n.notice_number,
               n.dispatch_date AS notice_date, n.due_date, n.status,
               n.created_at,
               c.name AS consumer_name, c.village
          FROM notices n
          LEFT JOIN raid_cases rc ON rc.case_id = n.case_id
          LEFT JOIN consumers c ON c.id = rc.consumer_id
         ORDER BY COALESCE(n.dispatch_date, n.created_at) DESC, n.id DESC
        """
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "Notices"
    headers = ["ID", "Case ID", "Type", "Notice #", "Notice Date",
               "Due Date", "Status", "Created",
               "Consumer", "Village"]
    _write_header(ws, headers)
    for r in rows:
        ws.append([
            r["id"], r["case_id"], r["notice_type"], r["notice_number"],
            r["notice_date"], r["due_date"], r["status"],
            r["created_at"],
            r["consumer_name"], r["village"],
        ])
    _autosize(ws)

    out = REPORTS_DIR / f"notices_{_stamp()}.xlsx"
    wb.save(out)
    return out


# ===================================================================
# Dashboard PDF (reportlab)
# ===================================================================
def dashboard_pdf() -> Path:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (Paragraph, SimpleDocTemplate, Spacer,
                                    Table, TableStyle)
    from reportlab.lib.enums import TA_CENTER

    out = REPORTS_DIR / f"dashboard_{_stamp()}.pdf"

    summary = fetch_all(
        """
        SELECT case_status,
               COUNT(*) AS count,
               COALESCE(SUM(total_assessment), 0) AS amount
          FROM raid_cases
         GROUP BY case_status
        """
    )
    today_pay = fetch_all(
        "SELECT COUNT(*) AS n, COALESCE(SUM(amount),0) AS amt "
        "FROM payments WHERE payment_date = ?",
        (date.today().isoformat(),),
    )[0]

    div_breakdown = fetch_all(
        """
        SELECT c.div_code, COUNT(*) AS cases,
               COALESCE(SUM(rc.total_assessment), 0) AS amount
          FROM raid_cases rc LEFT JOIN consumers c ON c.id = rc.consumer_id
         GROUP BY c.div_code
         ORDER BY cases DESC
        """
    )

    doc = SimpleDocTemplate(
        str(out), pagesize=A4,
        leftMargin=36, rightMargin=36, topMargin=36, bottomMargin=36,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "title", parent=styles["Title"], alignment=TA_CENTER,
        fontSize=16, spaceAfter=8,
    )
    sub_style = ParagraphStyle(
        "sub", parent=styles["Heading2"], alignment=TA_CENTER,
        fontSize=11, textColor=colors.grey, spaceAfter=14,
    )

    story: list[Any] = [
        Paragraph("Raid Management — Daily Dashboard", title_style),
        Paragraph(f"Generated: {datetime.now():%Y-%m-%d %H:%M}", sub_style),
    ]

    # Summary table
    rows = [["Status", "Cases", "Amount (₹)"]]
    total_cases = 0
    total_amt = 0.0
    for r in summary:
        rows.append([r["case_status"] or "—", r["count"],
                     f"{safe_float(r['amount']):,.2f}"])
        total_cases += int(r["count"])
        total_amt += safe_float(r["amount"])
    rows.append(["TOTAL", total_cases, f"{total_amt:,.2f}"])

    story.append(Paragraph("<b>Cases by Status</b>", styles["Heading2"]))
    t = Table(rows, hAlign="LEFT")
    t.setStyle(TableStyle([
        ("BACKGROUND",  (0, 0), (-1, 0), colors.HexColor("#305496")),
        ("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
        ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BACKGROUND",  (0, -1), (-1, -1), colors.HexColor("#FFE699")),
        ("FONTNAME",    (0, -1), (-1, -1), "Helvetica-Bold"),
        ("GRID",        (0, 0), (-1, -1), 0.5, colors.grey),
        ("ALIGN",       (1, 0), (-1, -1), "RIGHT"),
        ("ALIGN",       (0, 0), (0, -1), "LEFT"),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(t)
    story.append(Spacer(1, 14))

    # Today's payments
    story.append(Paragraph("<b>Today's Payments</b>", styles["Heading2"]))
    rows2 = [
        ["Count", today_pay["n"] or 0],
        ["Amount (₹)", f"{safe_float(today_pay['amt']):,.2f}"],
    ]
    t2 = Table(rows2, hAlign="LEFT", colWidths=[140, 140])
    t2.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
    ]))
    story.append(t2)
    story.append(Spacer(1, 14))

    # Division breakdown
    if div_breakdown:
        story.append(Paragraph("<b>By Division</b>", styles["Heading2"]))
        rows3 = [["Division", "Cases", "Amount (₹)"]]
        for r in div_breakdown:
            rows3.append([r["div_code"] or "—", r["cases"],
                         f"{safe_float(r['amount']):,.2f}"])
        t3 = Table(rows3, hAlign="LEFT")
        t3.setStyle(TableStyle([
            ("BACKGROUND",  (0, 0), (-1, 0), colors.HexColor("#305496")),
            ("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
            ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID",        (0, 0), (-1, -1), 0.5, colors.grey),
            ("ALIGN",       (1, 0), (-1, -1), "RIGHT"),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ]))
        story.append(t3)

    doc.build(story)
    log.info("dashboard_pdf → %s", out.name)
    return out


# ===================================================================
# Listing helper
# ===================================================================
def list_reports() -> list[dict]:
    items = []
    for p in sorted(REPORTS_DIR.glob("*.*"),
                    key=lambda x: x.stat().st_mtime, reverse=True):
        st = p.stat()
        items.append({
            "name": p.name,
            "size": st.st_size,
            "size_kb": round(st.st_size / 1024, 1),
            "modified": datetime.fromtimestamp(st.st_mtime)
                                 .isoformat(timespec="seconds"),
        })
    return items
