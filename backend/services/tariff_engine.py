"""
Tariff timeline engine (PR1).

Owns the `tariff_rates` table — a timeline-aware rate schedule that
extends the legacy `rate_master` table with:

* per-row effective_from / effective_to dates
* condition_load (load-band qualifier, e.g. "domestic", "industrial")
* slab_name (human label like "First 100 units")
* rebate / meter_rent (charge components)

This module is a *new, isolated surface*. PR1 deliberately does NOT
wire it into routes, the calculator, the notice generator, the LFHD
flow, or the frontend. Those integrations belong to PR2/PR3/PR4.

Public API
----------
* import_schedule(xlsx_path, ...)            — load workbook into tariff_rates
* find_applicable_rate(category, units, ...) — pick the best matching row
* get_schedule(schedule_name=None)           — read rows back (SELECT *)
* detect_overlaps(category, ...)             — find conflicting active rows
* build_sample_workbook(path)                — write a sample xlsx

Internals
---------
* _normalize_header_key(s)  — punctuation/case stripper, preserves Devanagari
* _canonical_header(s)      — synonym-aware canonical column name
* COLUMN_SYNONYMS           — header form -> canonical column
* _RATE_INSERT_SQL          — 19-placeholder INSERT (13 base + 6 PR1)
* _is_blank_rate_row(row)   — drop rows with no material content
* _SAMPLE_HEADERS / _SAMPLE_ROWS — embedded sample data
* _match / _specificity     — eligibility + tie-breaker for find_applicable_rate
* _ranges_overlap           — generic [a..b] x [c..d] intersection (None=open)
"""
from __future__ import annotations

import logging
import sqlite3
from contextlib import contextmanager
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable, Iterator, Optional

log = logging.getLogger(__name__)


# =====================================================================
# Header normalization
# =====================================================================
# Characters stripped to nothing during header normalization.
# Whitespace + common ASCII punctuation/separators. Devanagari, other
# Unicode letters, and digits are preserved.
_STRIP_CHARS = frozenset(" \t\n\r\v\f_.-/\\()%+,;:")


def _normalize_header_key(value: Any) -> str:
    """Lowercase ASCII letters, drop punctuation/whitespace, preserve Devanagari.

    Examples
    --------
    >>> _normalize_header_key("Effective From")
    'effectivefrom'
    >>> _normalize_header_key("CATEGORY_CODE")
    'categorycode'
    >>> _normalize_header_key("श्रेणी")
    'श्रेणी'
    """
    if value is None:
        return ""
    s = str(value).strip()
    out: list[str] = []
    for ch in s:
        if ch in _STRIP_CHARS:
            continue
        # Lowercase Latin letters; everything else (digits, Devanagari, etc.)
        # is preserved verbatim.
        if "A" <= ch <= "Z":
            out.append(chr(ord(ch) + 32))
        else:
            out.append(ch)
    return "".join(out)


# Map normalized header form -> canonical tariff_rates column name.
COLUMN_SYNONYMS: dict[str, str] = {
    # category
    "category":              "category",
    "categorycode":          "category",
    "consumercategory":      "category",
    "श्रेणी":                "category",
    # slab boundaries
    "slabstart":             "slab_start",
    "slabfrom":              "slab_start",
    "fromunits":             "slab_start",
    "unitsfrom":             "slab_start",
    "slabend":               "slab_end",
    "slabto":                "slab_end",
    "tounits":               "slab_end",
    "unitsto":               "slab_end",
    # rate
    "rate":                  "rate_per_unit",
    "rateperunit":           "rate_per_unit",
    "energyrate":            "rate_per_unit",
    "tariff":                "rate_per_unit",
    "tariffrate":            "rate_per_unit",
    # fixed charge
    "fixed":                 "fixed_charge",
    "fixedcharge":           "fixed_charge",
    "fixedcharges":          "fixed_charge",
    # duty
    "duty":                  "duty_percent",
    "dutypercent":           "duty_percent",
    "ed":                    "duty_percent",
    "edpercent":             "duty_percent",
    "electricityduty":       "duty_percent",
    # condition (free text)
    "condition":             "condition",
    "conditiontext":         "condition",
    # === PR1 NEW COLUMNS ===
    "conditionload":         "condition_load",
    "loadcondition":         "condition_load",
    "loadband":              "condition_load",
    "slabname":              "slab_name",
    "slablabel":             "slab_name",
    "rebate":                "rebate",
    "rebateamount":          "rebate",
    "meterrent":             "meter_rent",
    "metercharge":           "meter_rent",
    "metercharges":          "meter_rent",
    "effectivefrom":         "effective_from",
    "validfrom":             "effective_from",
    "fromdate":              "effective_from",
    "wef":                   "effective_from",
    "effectiveto":           "effective_to",
    "validto":               "effective_to",
    "todate":                "effective_to",
    "validtill":             "effective_to",
    # schedule-level metadata
    "schedulename":          "schedule_name",
    "scheduleeffectivefrom": "schedule_effective_from",
    "scheduleeffectiveto":   "schedule_effective_to",
    # bookkeeping
    "status":                "status",
    "notes":                 "notes",
    "remarks":               "notes",
    "source":                "source",
}


def _canonical_header(value: Any) -> Optional[str]:
    """Return the canonical column name for a header cell, or None."""
    return COLUMN_SYNONYMS.get(_normalize_header_key(value))


# =====================================================================
# INSERT SQL (13 base + 6 PR1 = 19 placeholders)
# =====================================================================
_RATE_INSERT_SQL = """
INSERT INTO tariff_rates (
    category, slab_start, slab_end, rate_per_unit, fixed_charge, duty_percent,
    condition, schedule_name, schedule_effective_from, schedule_effective_to,
    status, source, notes,
    condition_load, slab_name, rebate, meter_rent, effective_from, effective_to
) VALUES (?,?,?,?,?,?, ?,?,?,?, ?,?,?, ?,?,?,?,?,?)
""".strip()


# Fields considered when deciding whether an imported row is "blank".
_RATE_FIELDS_FOR_BLANK: tuple[str, ...] = (
    "category",
    "slab_start", "slab_end",
    "rate_per_unit", "fixed_charge", "duty_percent",
    "condition",
    "condition_load", "slab_name", "rebate", "meter_rent",
)


def _is_blank_rate_row(row: dict) -> bool:
    """True if every material field is None or empty whitespace."""
    if not row:
        return True
    for k in _RATE_FIELDS_FOR_BLANK:
        v = row.get(k)
        if v is None:
            continue
        if isinstance(v, str) and v.strip() == "":
            continue
        return False
    return True


# =====================================================================
# Type coercion helpers
# =====================================================================
def _coerce_text(v: Any) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _coerce_int(v: Any) -> Optional[int]:
    if v is None or v == "":
        return None
    try:
        return int(float(str(v).replace(",", "").strip()))
    except (TypeError, ValueError):
        return None


def _coerce_float(v: Any) -> Optional[float]:
    if v is None or v == "":
        return None
    try:
        return float(str(v).replace(",", "").strip())
    except (TypeError, ValueError):
        return None


def _coerce_date(v: Any) -> Optional[str]:
    """Best-effort -> ISO yyyy-mm-dd string."""
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d",
                "%d-%b-%Y", "%d %b %Y", "%d.%m.%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(s.split("T")[0]).date().isoformat()
    except Exception:  # noqa: BLE001
        return None


# =====================================================================
# Connection plumbing (works inside Flask request OR standalone)
# =====================================================================
def _dict_factory(cursor: sqlite3.Cursor, row: tuple) -> dict[str, Any]:
    return {col[0]: row[idx] for idx, col in enumerate(cursor.description)}


@contextmanager
def _conn_ctx(conn: Optional[sqlite3.Connection] = None) -> Iterator[sqlite3.Connection]:
    """Yield a sqlite3 connection.

    If `conn` is provided, it is used as-is (caller manages its lifecycle).
    Otherwise a fresh standalone connection is opened against config.DB_PATH.
    """
    if conn is not None:
        yield conn
        return
    from ..config import DB_PATH
    own = sqlite3.connect(str(DB_PATH), timeout=30.0)
    own.row_factory = _dict_factory
    try:
        yield own
        own.commit()
    finally:
        own.close()


# =====================================================================
# Sample workbook (used by tests + onboarding)
# =====================================================================
_SAMPLE_HEADERS: tuple[str, ...] = (
    "CATEGORY_CODE",
    "SLAB_START",
    "SLAB_END",
    "RATE_PER_UNIT",
    "FIXED_CHARGE",
    "DUTY_PERCENT",
    "CONDITION",
    "CONDITION_LOAD",
    "SLAB_NAME",
    "REBATE",
    "METER_RENT",
    "EFFECTIVE_FROM",
    "EFFECTIVE_TO",
)

# Each row is a tuple aligned 1-1 with _SAMPLE_HEADERS.
_SAMPLE_ROWS: tuple[tuple, ...] = (
    ("LMV-1", 0,   100,  5.50, 110.0, 5.0, "Domestic - First slab",
        "domestic", "First 100 units", 0.0,  20.0, "2025-04-01", "2026-03-31"),
    ("LMV-1", 101, 200,  6.00, 110.0, 5.0, "Domestic - Second slab",
        "domestic", "Next 100 units", 0.0,  20.0, "2025-04-01", "2026-03-31"),
    ("LMV-1", 201, None, 6.50, 110.0, 5.0, "Domestic - Above 200",
        "domestic", "Above 200 units", 0.0, 20.0, "2025-04-01", "2026-03-31"),
    ("LMV-2", 0,   None, 7.25, 150.0, 7.5, "Commercial flat rate",
        "commercial", "Flat", 0.0, 30.0, "2025-04-01", "2026-03-31"),
    ("LMV-6", 0,   None, 6.75, 200.0, 7.5, "Industrial",
        "industrial", "Flat", 50.0, 50.0, "2025-04-01", "2026-03-31"),
)


def build_sample_workbook(path: Any) -> Path:
    """Write a minimal tariff schedule workbook usable by import_schedule()."""
    import openpyxl
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "tariff"
    ws.append(list(_SAMPLE_HEADERS))
    for row in _SAMPLE_ROWS:
        ws.append(list(row))
    wb.save(str(p))
    return p


# =====================================================================
# Importer
# =====================================================================
def _read_workbook_records(xlsx_path: Any) -> tuple[list[Optional[str]], list[dict]]:
    """Return (canonical_header_list, list_of_dict_records)."""
    import openpyxl
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    ws = wb.active
    iter_rows = ws.iter_rows(values_only=True)
    raw_headers = next(iter_rows, None) or ()
    canonical = [_canonical_header(h) for h in raw_headers]
    records: list[dict] = []
    for raw in iter_rows:
        rec: dict[str, Any] = {}
        for col_name, val in zip(canonical, raw):
            if col_name is None:
                continue
            rec[col_name] = val
        records.append(rec)
    return canonical, records


def _import_records(records: Iterable[dict],
                    schedule_name: Optional[str],
                    schedule_effective_from: Optional[str],
                    schedule_effective_to: Optional[str],
                    source: Optional[str],
                    conn: sqlite3.Connection) -> dict:
    inserted = 0
    skipped_blank = 0
    sched_eff_from = _coerce_date(schedule_effective_from)
    sched_eff_to = _coerce_date(schedule_effective_to)
    for rec in records:
        if _is_blank_rate_row(rec):
            skipped_blank += 1
            continue
        params = (
            _coerce_text(rec.get("category")),
            _coerce_int(rec.get("slab_start")),
            _coerce_int(rec.get("slab_end")),
            _coerce_float(rec.get("rate_per_unit")),
            _coerce_float(rec.get("fixed_charge")),
            _coerce_float(rec.get("duty_percent")),
            _coerce_text(rec.get("condition")),
            _coerce_text(rec.get("schedule_name")) or schedule_name,
            _coerce_date(rec.get("schedule_effective_from")) or sched_eff_from,
            _coerce_date(rec.get("schedule_effective_to")) or sched_eff_to,
            _coerce_text(rec.get("status")) or "active",
            _coerce_text(rec.get("source")) or source,
            _coerce_text(rec.get("notes")),
            # PR1 columns
            _coerce_text(rec.get("condition_load")),
            _coerce_text(rec.get("slab_name")),
            _coerce_float(rec.get("rebate")),
            _coerce_float(rec.get("meter_rent")),
            _coerce_date(rec.get("effective_from")),
            _coerce_date(rec.get("effective_to")),
        )
        conn.execute(_RATE_INSERT_SQL, params)
        inserted += 1
    return {
        "schedule_name": schedule_name,
        "inserted": inserted,
        "skipped_blank": skipped_blank,
    }


def import_schedule(xlsx_path: Any,
                    schedule_name: Optional[str] = None,
                    schedule_effective_from: Optional[str] = None,
                    schedule_effective_to: Optional[str] = None,
                    source: Optional[str] = None,
                    conn: Optional[sqlite3.Connection] = None) -> dict:
    """Import a tariff workbook into the tariff_rates table.

    The workbook's first row is treated as headers; each remaining row
    becomes one tariff_rates record. Headers are normalized via
    _normalize_header_key + COLUMN_SYNONYMS, so column ordering and
    casing in the source file do not matter.
    """
    p = Path(str(xlsx_path))
    sched = schedule_name or p.stem
    src = source or str(p)
    _, records = _read_workbook_records(p)
    with _conn_ctx(conn) as c:
        result = _import_records(
            records,
            schedule_name=sched,
            schedule_effective_from=schedule_effective_from,
            schedule_effective_to=schedule_effective_to,
            source=src,
            conn=c,
        )
        if conn is None:
            c.commit()
    return result


# =====================================================================
# Read-back: get_schedule  (SELECT * so PR1 columns travel with the row)
# =====================================================================
def get_schedule(schedule_name: Optional[str] = None,
                 conn: Optional[sqlite3.Connection] = None) -> list[dict]:
    """Return tariff_rates rows (optionally filtered by schedule_name).

    Uses ``SELECT *`` so that any new column added by future migrations is
    automatically returned to callers without a code change here.
    """
    with _conn_ctx(conn) as c:
        prev_factory = c.row_factory
        c.row_factory = _dict_factory
        try:
            if schedule_name is None:
                rows = c.execute(
                    "SELECT * FROM tariff_rates ORDER BY id"
                ).fetchall()
            else:
                rows = c.execute(
                    "SELECT * FROM tariff_rates WHERE schedule_name = ? "
                    "ORDER BY id",
                    (schedule_name,),
                ).fetchall()
        finally:
            c.row_factory = prev_factory
    return [dict(r) for r in rows]


# =====================================================================
# Range overlap utility (used by detect_overlaps + find_applicable_rate)
# =====================================================================
def _ranges_overlap(a_start: Any, a_end: Any,
                    b_start: Any, b_end: Any) -> bool:
    """True iff [a_start..a_end] intersects [b_start..b_end].

    None on any boundary means "open" on that side. Comparison is done
    with the natural ordering of the supplied values, so this works for
    both ints (slab boundaries) and ISO yyyy-mm-dd strings (dates).
    """
    if a_end is not None and b_start is not None and a_end < b_start:
        return False
    if b_end is not None and a_start is not None and b_end < a_start:
        return False
    return True


# =====================================================================
# find_applicable_rate
# =====================================================================
def _to_iso(v: Any) -> Optional[str]:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return _coerce_date(v)


def _match(row: dict,
           units: float,
           as_of: Optional[str],
           condition_load: Optional[str]) -> bool:
    # Status
    status = (row.get("status") or "active").strip().lower()
    if status and status != "active":
        return False
    # Slab boundaries
    start = row.get("slab_start")
    end = row.get("slab_end")
    if start is not None and units < float(start):
        return False
    if end is not None and units > float(end):
        return False
    # Date filter — per-row dates take priority; fall back to schedule-level.
    if as_of:
        eff_from = row.get("effective_from") or row.get("schedule_effective_from")
        eff_to = row.get("effective_to") or row.get("schedule_effective_to")
        if eff_from and as_of < eff_from:
            return False
        if eff_to and as_of > eff_to:
            return False
    # condition_load filter — only enforce when both row and arg are present.
    row_cl = row.get("condition_load")
    if condition_load and row_cl:
        if str(row_cl).strip().lower() != str(condition_load).strip().lower():
            return False
    return True


def _specificity(row: dict) -> int:
    """Higher = more specific. Per-row dates dominate (+16) per PR1 spec."""
    score = 0
    if row.get("slab_end") is not None:
        score += 4
    if row.get("slab_start") is not None and (row.get("slab_start") or 0) > 0:
        score += 2
    if row.get("condition_load"):
        score += 8
    if row.get("effective_from") or row.get("effective_to"):
        score += 16
    if row.get("condition"):
        score += 1
    return score


def find_applicable_rate(category: str,
                         units_consumed: float,
                         as_of_date: Any = None,
                         condition_load: Optional[str] = None,
                         conn: Optional[sqlite3.Connection] = None) -> Optional[dict]:
    """Pick the tariff_rates row that best applies.

    Returns the chosen row as a dict (with all PR1 columns including
    schedule_effective_from/to AND per-row effective_from/to), or None.
    """
    as_of = _to_iso(as_of_date) if as_of_date else None
    with _conn_ctx(conn) as c:
        prev_factory = c.row_factory
        c.row_factory = _dict_factory
        try:
            raw = c.execute(
                "SELECT * FROM tariff_rates WHERE category = ?",
                (category,),
            ).fetchall()
        finally:
            c.row_factory = prev_factory
    rows = [dict(r) for r in raw]
    eligible = [r for r in rows
                if _match(r, float(units_consumed), as_of, condition_load)]
    if not eligible:
        return None
    eligible.sort(key=_specificity, reverse=True)
    return eligible[0]


# =====================================================================
# Overlap detection
# =====================================================================
def detect_overlaps(category: str,
                    slab_start: Any = None,
                    slab_end: Any = None,
                    effective_from: Any = None,
                    effective_to: Any = None,
                    condition_load: Optional[str] = None,
                    exclude_id: Optional[int] = None,
                    conn: Optional[sqlite3.Connection] = None) -> list[dict]:
    """Return active tariff_rates rows that conflict with the proposed range.

    Two rows are considered conflicting (overlapping) iff ALL hold:

      * same category (literal match)
      * condition_load values are *compatible* — both NULL, OR one NULL
        (NULL acts as a wildcard), OR equal (case-insensitive)
      * slab ranges intersect (NULL slab_start treated as 0,
        NULL slab_end treated as +infinity)
      * effective date windows intersect, where the window is
        ``effective_from..effective_to`` if either is set, else falling
        back to ``schedule_effective_from..schedule_effective_to``;
        NULL boundaries are open on that side
      * the existing row has status='active' (or NULL, treated as active)

    The proposed window follows the same NULL-boundary convention.
    `exclude_id` lets callers skip a particular row id when checking
    an already-stored row against its peers.
    """
    a_slab_start = _coerce_int(slab_start)
    if a_slab_start is None:
        a_slab_start = 0
    a_slab_end = _coerce_int(slab_end)  # may stay None (open)
    a_eff_from = _to_iso(effective_from)
    a_eff_to = _to_iso(effective_to)
    a_cl_raw = condition_load.strip() if isinstance(condition_load, str) else condition_load
    a_cl = a_cl_raw.lower() if isinstance(a_cl_raw, str) and a_cl_raw else None

    with _conn_ctx(conn) as c:
        prev_factory = c.row_factory
        c.row_factory = _dict_factory
        try:
            rows = c.execute(
                "SELECT * FROM tariff_rates WHERE category = ?",
                (category,),
            ).fetchall()
        finally:
            c.row_factory = prev_factory

    overlaps: list[dict] = []
    for r in rows:
        rid = r.get("id")
        if exclude_id is not None and rid == exclude_id:
            continue
        status = (r.get("status") or "active").strip().lower()
        if status != "active":
            continue
        # condition_load compatibility — NULL is a wildcard.
        b_cl_raw = r.get("condition_load")
        b_cl = (str(b_cl_raw).strip().lower()
                if b_cl_raw is not None and str(b_cl_raw).strip() else None)
        if a_cl is not None and b_cl is not None and a_cl != b_cl:
            continue
        # slab overlap
        b_slab_start = r.get("slab_start") if r.get("slab_start") is not None else 0
        b_slab_end = r.get("slab_end")
        if not _ranges_overlap(a_slab_start, a_slab_end, b_slab_start, b_slab_end):
            continue
        # date overlap — per-row dates take priority over schedule-level
        b_eff_from = r.get("effective_from") or r.get("schedule_effective_from")
        b_eff_to = r.get("effective_to") or r.get("schedule_effective_to")
        if not _ranges_overlap(a_eff_from, a_eff_to, b_eff_from, b_eff_to):
            continue
        overlaps.append(dict(r))
    return overlaps
