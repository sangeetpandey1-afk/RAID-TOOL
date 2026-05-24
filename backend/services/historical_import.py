"""
Historical offense Excel ingestion (PR2).

Purpose
-------
PR4's Offense Verification UI needs to display 15 fields per matched
historical offense row:

    Notice No., Div. No., Date, Name, Father Name, Use Name,
    User Father Name, Address, Sub Station, Assessment, Old AC No.,
    New Account Number, Category, Irregularity, Paid/Unpaid

The legacy `historical_cases` table only has 9 of those (id, div_no,
name, father_name, village, account_id, case_date, assessment_amount,
fir_number, section). PR2's lightweight migration in database.py adds
the remaining 10 nullable columns. THIS module reads the new historical
Excel format and writes those columns.

Why a separate file (not the existing importer.py)?
----------------------------------------------------
* importer.py is pandas-based; the sandbox + minimal-deps install path
  may not have pandas. This module uses ONLY openpyxl (already a hard
  dep of the tariff engine).
* Keeping a dedicated module avoids touching the pandas importer that
  the existing /api/import_all_master_data flow depends on. The legacy
  importer keeps its current behaviour for backward-compat; PR4 routes
  use this new function.

Public API
----------
* import_historical_workbook(xlsx_path, source=None, conn=None) -> dict
* _read_historical_records(xlsx_path) -> (headers, list_of_dicts)
* HISTORICAL_COLUMN_SYNONYMS

The synonym registry follows the same convention as
tariff_engine.COLUMN_SYNONYMS — keys are normalized (whitespace +
punctuation stripped, ASCII lowercased, Devanagari preserved).

Compatibility
-------------
* Does NOT modify importer.py.
* Does NOT modify the historical_cases legacy columns; only writes the
  new PR2 columns + the existing ones.
* Idempotent on (notice_no, account_id, case_date) — running the same
  workbook twice does not insert duplicates.
"""
from __future__ import annotations

import logging
import sqlite3
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable, Optional

from . import tariff_engine as te
from .tariff_engine import (
    _coerce_date,
    _coerce_float,
    _coerce_text,
    _conn_ctx,
    _normalize_header_key,
)

log = logging.getLogger(__name__)


# =====================================================================
# Header synonyms (normalized form -> canonical historical_cases column)
# =====================================================================
HISTORICAL_COLUMN_SYNONYMS: dict[str, str] = {
    # Notice number
    "noticeno":            "notice_no",
    "noticenumber":        "notice_no",
    "noticenum":           "notice_no",
    "नोटिसक्रमांक":         "notice_no",
    # Division
    "divno":               "div_no",
    "divisionno":          "div_no",
    "divcode":             "div_no",
    "divisioncode":        "div_no",
    # Date
    "date":                "case_date",
    "casedate":            "case_date",
    "noticedate":          "case_date",
    "raiddate":            "case_date",
    "inspectiondate":      "case_date",
    "तिथि":                "case_date",
    "दिनांक":              "case_date",
    # Name (registered consumer)
    "name":                "name",
    "consumername":        "name",
    "customername":        "name",
    "नाम":                 "name",
    # Father name
    "fathername":          "father_name",
    "father":              "father_name",
    "fathersname":         "father_name",
    "पिता":                "father_name",
    # Use Name (user found at premises)
    "username":            "use_name",
    "usename":             "use_name",
    "userfoundname":       "use_name",
    "userfoundatpremises": "use_name",
    # User Father Name
    "userfathername":      "user_father_name",
    "userfather":          "user_father_name",
    "usefathername":       "user_father_name",
    "usersfather":         "user_father_name",
    # Address
    "address":             "address",
    "fulladdress":         "address",
    "premises":            "address",
    "पता":                 "address",
    # Village (legacy column, kept for back-compat)
    "village":             "village",
    "ग्राम":                "village",
    # Sub-station
    "substation":          "sub_substation",
    "subsubstation":       "sub_substation",
    "ss":                  "sub_substation",
    "supplystation":       "sub_substation",
    # Assessment amount
    "assessment":          "assessment_amount",
    "assessmentamount":    "assessment_amount",
    "amount":              "assessment_amount",
    "totalassessment":     "assessment_amount",
    "राशि":                "assessment_amount",
    # Old account id
    "oldacno":             "old_account_id",
    "oldaccount":          "old_account_id",
    "oldaccountnumber":    "old_account_id",
    "oldaccno":            "old_account_id",
    "oldid":               "old_account_id",
    # New account id (also the canonical "account_id" used by indexes)
    "newacno":             "new_account_id",
    "newaccount":          "new_account_id",
    "newaccountnumber":    "new_account_id",
    "newaccno":            "new_account_id",
    "accountid":           "new_account_id",
    "accountno":           "new_account_id",
    "acno":                "new_account_id",
    "kno":                 "new_account_id",
    # Category (LMV-1, LMV-2 etc.)
    "category":            "category",
    "tariffcategory":      "category",
    "ratecategory":        "category",
    "lmv":                 "category",
    "श्रेणी":              "category",
    # Irregularity / theft type
    "irregularity":        "irregularity",
    "theft":               "irregularity",
    "naturefviolation":    "irregularity",
    "violation":           "irregularity",
    "अनियमितता":           "irregularity",
    # Paid / unpaid
    "paid":                "paid_status",
    "paidunpaid":          "paid_status",
    "paymentstatus":       "paid_status",
    "paystatus":           "paid_status",
    "status":              "paid_status",
    # Other useful legacy columns
    "fir":                 "fir_number",
    "firno":               "fir_number",
    "firnumber":           "fir_number",
    "section":             "section",
    "dhara":               "section",
    "धारा":                "section",
}


def _canonical_historical_header(value: Any) -> Optional[str]:
    return HISTORICAL_COLUMN_SYNONYMS.get(_normalize_header_key(value))


# =====================================================================
# Insert SQL — covers legacy + PR2 extension columns
# =====================================================================
_HIST_INSERT_SQL = """
INSERT INTO historical_cases (
    div_no, name, father_name, village, account_id,
    case_date, assessment_amount, fir_number, section, source,
    notice_no, address, use_name, user_father_name, sub_substation,
    old_account_id, new_account_id, category, irregularity, paid_status
) VALUES (
    ?,?,?,?,?, ?,?,?,?,?, ?,?,?,?,?, ?,?,?,?,?
)
""".strip()

# Used for idempotency check before insert.
_HIST_DEDUPE_SQL = """
SELECT id FROM historical_cases
WHERE COALESCE(notice_no,'') = COALESCE(?,'')
  AND COALESCE(account_id,'') = COALESCE(?,'')
  AND COALESCE(case_date,'')  = COALESCE(?,'')
LIMIT 1
""".strip()


# Material fields used to decide whether a record is "blank" (skip it).
_HIST_FIELDS_FOR_BLANK: tuple[str, ...] = (
    "notice_no", "div_no", "name", "father_name",
    "account_id", "new_account_id", "old_account_id",
    "case_date", "assessment_amount",
    "address", "use_name", "user_father_name",
    "sub_substation", "category", "irregularity", "paid_status",
)


def _is_blank_historical_row(rec: dict) -> bool:
    if not rec:
        return True
    for k in _HIST_FIELDS_FOR_BLANK:
        v = rec.get(k)
        if v is None:
            continue
        if isinstance(v, str) and v.strip() == "":
            continue
        return False
    return True


# =====================================================================
# Read workbook
# =====================================================================
def _read_historical_records(xlsx_path: Any) -> tuple[list, list[dict]]:
    """
    Return (canonical_header_list, list_of_dict_records).

    The first row of the active sheet is treated as the header row.
    """
    import openpyxl
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    ws = wb.active
    iter_rows = ws.iter_rows(values_only=True)
    raw_headers = next(iter_rows, None) or ()
    canonical = [_canonical_historical_header(h) for h in raw_headers]
    records: list[dict] = []
    for raw in iter_rows:
        rec: dict[str, Any] = {}
        for col_name, val in zip(canonical, raw):
            if col_name is None:
                continue
            # Last-write-wins if a column appears twice — that's fine for
            # the malformed-Excel cases we observed in the wild.
            rec[col_name] = val
        records.append(rec)
    return canonical, records


# =====================================================================
# Type coercion helpers (account-number specific)
# =====================================================================
def _coerce_account(v: Any) -> Optional[str]:
    """Account numbers — strip non-alphanumeric, uppercase."""
    if v is None:
        return None
    import re
    s = re.sub(r"[^A-Za-z0-9]", "", str(v)).upper()
    return s or None


def _coerce_paid_status(v: Any) -> Optional[str]:
    """Normalize paid/unpaid values to lowercase 'paid' / 'unpaid' / raw."""
    if v is None:
        return None
    s = str(v).strip().lower()
    if not s:
        return None
    if s in ("y", "yes", "p", "paid", "✓", "tick", "complete", "completed",
             "cleared", "settled"):
        return "paid"
    if s in ("n", "no", "u", "unpaid", "due", "pending", "outstanding",
             "x", "✗"):
        return "unpaid"
    return s  # preserve "partial", custom strings, etc.


# =====================================================================
# Public: import_historical_workbook
# =====================================================================
def import_historical_workbook(xlsx_path: Any,
                               source: Optional[str] = None,
                               conn: Optional[sqlite3.Connection] = None
                               ) -> dict:
    """
    Read a historical-offense Excel and insert its rows into
    historical_cases. Idempotent on (notice_no, account_id, case_date).

    Returns:
        {
            "ok": True/False,
            "file": "...",
            "headers_seen": [...],
            "headers_mapped": {raw -> canonical},
            "inserted": N,
            "skipped_blank": N,
            "skipped_duplicate": N,
            "errors": [{"row": idx, "error": "..."}, ...],
        }
    """
    p = Path(str(xlsx_path))
    if not p.exists():
        return {"ok": False, "error": f"File not found: {p}"}

    src = source or str(p.resolve())
    inserted = 0
    skipped_blank = 0
    skipped_dup = 0
    errors: list[dict] = []
    try:
        canonical, records = _read_historical_records(p)
    except Exception as e:  # noqa: BLE001
        return {
            "ok": False,
            "error": f"Could not read workbook: {type(e).__name__}: {e}",
            "file": str(p),
        }

    headers_mapped = {raw: c for raw, c in zip(_first_row_strs(p), canonical)
                      if c is not None}
    headers_seen = list(_first_row_strs(p))

    with _conn_ctx(conn) as c:
        prev_factory = c.row_factory
        c.row_factory = te._dict_factory
        try:
            for idx, rec in enumerate(records):
                if _is_blank_historical_row(rec):
                    skipped_blank += 1
                    continue
                try:
                    # Prefer new_account_id if both old and new appear in the row
                    # (the legacy account_id column gets the most useful value).
                    new_acc = _coerce_account(rec.get("new_account_id"))
                    old_acc = _coerce_account(rec.get("old_account_id"))
                    primary_acc = new_acc or old_acc

                    notice_no = _coerce_text(rec.get("notice_no"))
                    case_date = _coerce_date(rec.get("case_date"))

                    # Idempotency — skip if (notice_no, account, date) seen
                    if c.execute(_HIST_DEDUPE_SQL,
                                 (notice_no, primary_acc, case_date)
                                 ).fetchone() is not None:
                        skipped_dup += 1
                        continue

                    params = (
                        _coerce_text(rec.get("div_no")),
                        _coerce_text(rec.get("name")),
                        _coerce_text(rec.get("father_name")),
                        _coerce_text(rec.get("village")),
                        primary_acc,
                        case_date,
                        _coerce_float(rec.get("assessment_amount")),
                        _coerce_text(rec.get("fir_number")),
                        _coerce_text(rec.get("section")),
                        src,
                        # PR2 extension columns
                        notice_no,
                        _coerce_text(rec.get("address")),
                        _coerce_text(rec.get("use_name")),
                        _coerce_text(rec.get("user_father_name")),
                        _coerce_text(rec.get("sub_substation")),
                        old_acc,
                        new_acc,
                        _coerce_text(rec.get("category")),
                        _coerce_text(rec.get("irregularity")),
                        _coerce_paid_status(rec.get("paid_status")),
                    )
                    c.execute(_HIST_INSERT_SQL, params)
                    inserted += 1
                except Exception as e:  # noqa: BLE001
                    errors.append({
                        "row": idx + 2,  # +2 = header row + 1-index
                        "error": f"{type(e).__name__}: {e}",
                    })
            if conn is None:
                c.commit()
        finally:
            c.row_factory = prev_factory

    return {
        "ok": True,
        "file": str(p),
        "headers_seen": headers_seen,
        "headers_mapped": headers_mapped,
        "rows_total": len(records),
        "inserted": inserted,
        "skipped_blank": skipped_blank,
        "skipped_duplicate": skipped_dup,
        "errors": errors,
    }


def _first_row_strs(xlsx_path: Any) -> list[str]:
    """Read just the header row's literal strings (for diagnostics)."""
    import openpyxl
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True, read_only=True)
    ws = wb.active
    out: list[str] = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        for v in row:
            out.append("" if v is None else str(v))
        break
    return out


# =====================================================================
# Sample workbook (for tests / onboarding) — uses the new 15-field layout
# =====================================================================
_HIST_SAMPLE_HEADERS: tuple[str, ...] = (
    "Notice No.",
    "Div. No.",
    "Date",
    "Name",
    "Father Name",
    "Use Name",
    "User Father Name",
    "Address",
    "Sub Station",
    "Assessment",
    "Old AC No.",
    "New Account Number",
    "Category",
    "Irregularity",
    "Paid/Unpaid",
)

_HIST_SAMPLE_ROWS: tuple[tuple, ...] = (
    ("N-2024-001", "DIV-12", "2024-05-12", "RAM KUMAR", "SHYAM LAL",
        "RAM KUMAR", "SHYAM LAL", "VPO RAMPUR", "SS-12",
        12500.0, "OLD123", "NEW456", "LMV-1",
        "Direct theft via hook", "unpaid"),
    ("N-2025-007", "DIV-12", "2025-02-03", "GEETA DEVI", "MOHAN LAL",
        "GEETA DEVI", "MOHAN LAL", "VPO MOHANGANJ", "SS-08",
        8400.0, "OLD200", "NEW501", "LMV-1",
        "Meter tampering", "paid"),
    ("N-2025-019", "DIV-13", "2025-08-15", "AJAY SINGH", "BIRENDRA SINGH",
        "AJAY SINGH", "BIRENDRA SINGH", "VPO AJAYPUR", "SS-15",
        21500.0, "", "NEW777", "LMV-2",
        "Commercial use on domestic", "unpaid"),
)


def build_sample_historical_workbook(path: Any) -> Path:
    """Write a minimal historical-offense workbook for tests/onboarding."""
    import openpyxl
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "historical"
    ws.append(list(_HIST_SAMPLE_HEADERS))
    for row in _HIST_SAMPLE_ROWS:
        ws.append(list(row))
    wb.save(str(p))
    return p
