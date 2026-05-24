"""
Generate SAMPLE Excel master-data files.

Why
---
Officers often ask: "kis column ke saath kaunsa heading lagana hai?"
The importer (backend/services/importer.py) accepts MANY synonyms for each
target field, but to make life easy this script writes 6 ready-to-inspect
sample workbooks into ``master_data/`` using the *canonical* column names
that match the user's existing Excel files (per project-spec.md).

Each sheet has 5–8 realistic rows so you can visually compare the structure
with your real production files before uploading.

Usage
-----
    python scripts/generate_sample_excels.py

Files produced (all .gitignore-allowlisted via ``!master_data/SAMPLE_*.xlsx``):

    master_data/SAMPLE_raid_master_data.xlsx     (consumers)
    master_data/SAMPLE_all_data.xlsx             (historical cases)
    master_data/SAMPLE_raid_excell_2526.xlsx     (current/active cases)
    master_data/SAMPLE_device_list.xlsx          (device master)
    master_data/SAMPLE_slab_rates.xlsx           (rate / tariff slabs)
    master_data/SAMPLE_account_mapping.xlsx      (old <-> new account map)
"""
from __future__ import annotations
from pathlib import Path
from datetime import date

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
MASTER = ROOT / "master_data"
MASTER.mkdir(parents=True, exist_ok=True)


# =====================================================================
# 1. CONSUMER MASTER  (raid_master_data.xlsx)
# =====================================================================
CONSUMERS = [
    {"ACCT_ID": "1234567890", "DIV_CODE": "UPL01", "NAME": "रामकुमार सिंह",
     "FATHER_NAME": "श्यामलाल सिंह", "ADDRESS": "मकान नं. 12, गांधी नगर",
     "VILLAGE": "रामपुर", "LANDMARK": "स्कूल के पास", "POST": "रामपुर",
     "PIN": "244901", "TEHSIL": "बिलासपुर", "DISTRICT": "रामपुर",
     "MOBILE_NO": "9876543210", "LOAD": 2.0, "LOAD_UNIT": "KW",
     "SUPPLY_TYPE": "1Phase", "CATEGORY": "LMV-1", "SUB_SUBSTATION": "33/11 KV रामपुर",
     "CON_STATUS": "Active", "SC_NO": "SC100234"},
    {"ACCT_ID": "1234567891", "DIV_CODE": "UPL01", "NAME": "गीता देवी",
     "FATHER_NAME": "रामनाथ", "ADDRESS": "वार्ड 5, मेन रोड",
     "VILLAGE": "मिलक", "LANDMARK": "मंदिर के सामने", "POST": "मिलक",
     "PIN": "244925", "TEHSIL": "मिलक", "DISTRICT": "रामपुर",
     "MOBILE_NO": "9876501234", "LOAD": 1.5, "LOAD_UNIT": "KW",
     "SUPPLY_TYPE": "1Phase", "CATEGORY": "LMV-1", "SUB_SUBSTATION": "33/11 KV मिलक",
     "CON_STATUS": "Active", "SC_NO": "SC100235"},
    {"ACCT_ID": "1234567892", "DIV_CODE": "UPL02", "NAME": "Mohit Sharma",
     "FATHER_NAME": "Suresh Sharma", "ADDRESS": "Shop No. 7, Bazar Road",
     "VILLAGE": "Suar", "LANDMARK": "Near Petrol Pump", "POST": "Suar",
     "PIN": "244925", "TEHSIL": "Suar", "DISTRICT": "Rampur",
     "MOBILE_NO": "9123456780", "LOAD": 5.0, "LOAD_UNIT": "KW",
     "SUPPLY_TYPE": "3Phase", "CATEGORY": "LMV-2", "SUB_SUBSTATION": "33/11 KV Suar",
     "CON_STATUS": "Active", "SC_NO": "SC100236"},
    {"ACCT_ID": "1234567893", "DIV_CODE": "UPL01", "NAME": "Hari Prakash",
     "FATHER_NAME": "Late Devraj", "ADDRESS": "Mohalla Banjaran",
     "VILLAGE": "Chamraua", "LANDMARK": "Old Well", "POST": "Chamraua",
     "PIN": "244925", "TEHSIL": "Chamraua", "DISTRICT": "Rampur",
     "MOBILE_NO": "9000111222", "LOAD": 7.5, "LOAD_UNIT": "KW",
     "SUPPLY_TYPE": "3Phase", "CATEGORY": "LMV-5", "SUB_SUBSTATION": "33/11 KV Chamraua",
     "CON_STATUS": "Disconnected", "SC_NO": "SC100237"},
    {"ACCT_ID": "1234567894", "DIV_CODE": "UPL03", "NAME": "अनिल कुमार",
     "FATHER_NAME": "बद्री प्रसाद", "ADDRESS": "ग्राम पंचायत कार्यालय के पीछे",
     "VILLAGE": "बिलासपुर", "LANDMARK": "पंचायत भवन", "POST": "बिलासपुर",
     "PIN": "244921", "TEHSIL": "बिलासपुर", "DISTRICT": "रामपुर",
     "MOBILE_NO": "8888777666", "LOAD": 3.0, "LOAD_UNIT": "KW",
     "SUPPLY_TYPE": "1Phase", "CATEGORY": "LMV-1", "SUB_SUBSTATION": "33/11 KV बिलासपुर",
     "CON_STATUS": "Active", "SC_NO": "SC100238"},
    {"ACCT_ID": "1234567895", "DIV_CODE": "UPL02", "NAME": "Sunita Yadav",
     "FATHER_NAME": "Ram Yadav", "ADDRESS": "House 22, Agriculture Colony",
     "VILLAGE": "Tanda", "LANDMARK": "Mandi Gate", "POST": "Tanda",
     "PIN": "244925", "TEHSIL": "Tanda", "DISTRICT": "Rampur",
     "MOBILE_NO": "7777888999", "LOAD": 10.0, "LOAD_UNIT": "KW",
     "SUPPLY_TYPE": "3Phase", "CATEGORY": "LMV-5", "SUB_SUBSTATION": "33/11 KV Tanda",
     "CON_STATUS": "Active", "SC_NO": "SC100239"},
    {"ACCT_ID": "1234567896", "DIV_CODE": "UPL01", "NAME": "विजय वर्मा",
     "FATHER_NAME": "रमेश वर्मा", "ADDRESS": "नई बस्ती, गली नं. 3",
     "VILLAGE": "रामपुर", "LANDMARK": "बस स्टैंड", "POST": "रामपुर",
     "PIN": "244901", "TEHSIL": "रामपुर", "DISTRICT": "रामपुर",
     "MOBILE_NO": "9111222333", "LOAD": 2.0, "LOAD_UNIT": "KW",
     "SUPPLY_TYPE": "1Phase", "CATEGORY": "LMV-1", "SUB_SUBSTATION": "33/11 KV रामपुर",
     "CON_STATUS": "Active", "SC_NO": "SC100240"},
]


# =====================================================================
# 2. HISTORICAL CASES  (ALL DATA.xlsx)
# =====================================================================
HISTORICAL = [
    {"div no": "UPL01", "Name": "रामकुमार सिंह", "father name": "श्यामलाल सिंह",
     "village": "रामपुर", "Account Id": "1234567890", "Date": date(2022, 5, 12),
     "assessment": 45000, "FIR": "FIR/2022/0123", "dhara": "135"},
    {"div no": "UPL01", "Name": "गीता देवी", "father name": "रामनाथ",
     "village": "मिलक", "Account Id": "1234567891", "Date": date(2021, 11, 3),
     "assessment": 18500, "FIR": "FIR/2021/0987", "dhara": "138"},
    {"div no": "UPL02", "Name": "Mohit Sharma", "father name": "Suresh Sharma",
     "village": "Suar", "Account Id": "1234567892", "Date": date(2023, 2, 18),
     "assessment": 87000, "FIR": "FIR/2023/0045", "dhara": "135"},
    {"div no": "UPL01", "Name": "Hari Prakash", "father name": "Late Devraj",
     "village": "Chamraua", "Account Id": "1234567893", "Date": date(2020, 7, 22),
     "assessment": 62000, "FIR": "", "dhara": "126"},
    {"div no": "UPL03", "Name": "अनिल कुमार", "father name": "बद्री प्रसाद",
     "village": "बिलासपुर", "Account Id": "1234567894", "Date": date(2024, 1, 8),
     "assessment": 32000, "FIR": "FIR/2024/0011", "dhara": "135"},
    {"div no": "UPL01", "Name": "रामकुमार सिंह", "father name": "श्यामलाल सिंह",
     "village": "रामपुर", "Account Id": "1234567890", "Date": date(2024, 9, 14),
     "assessment": 55000, "FIR": "FIR/2024/0567", "dhara": "135"},
    {"div no": "UPL02", "Name": "Sunita Yadav", "father name": "Ram Yadav",
     "village": "Tanda", "Account Id": "1234567895", "Date": date(2023, 6, 30),
     "assessment": 120000, "FIR": "FIR/2023/0301", "dhara": "135"},
]


# =====================================================================
# 3. CURRENT (active) CASES  (raid excell 2526.xlsx)
# =====================================================================
CURRENT = [
    {"ONLINE NO": "ON2526001", "div no": "UPL01", "Name": "रामकुमार सिंह",
     "father name": "श्यामलाल सिंह", "village": "रामपुर",
     "connection no": "1234567890", "inspection_date": date(2025, 11, 5),
     "section": "135", "assessment_total": 95000,
     "notice_status": "dispatched", "payment_status": "pending"},
    {"ONLINE NO": "ON2526002", "div no": "UPL01", "Name": "विजय वर्मा",
     "father name": "रमेश वर्मा", "village": "रामपुर",
     "connection no": "1234567896", "inspection_date": date(2025, 12, 18),
     "section": "138", "assessment_total": 22000,
     "notice_status": "pending", "payment_status": "pending"},
    {"ONLINE NO": "ON2526003", "div no": "UPL02", "Name": "Mohit Sharma",
     "father name": "Suresh Sharma", "village": "Suar",
     "connection no": "1234567892", "inspection_date": date(2026, 1, 4),
     "section": "135", "assessment_total": 178000,
     "notice_status": "dispatched", "payment_status": "partial"},
    {"ONLINE NO": "ON2526004", "div no": "UPL03", "Name": "अनिल कुमार",
     "father name": "बद्री प्रसाद", "village": "बिलासपुर",
     "connection no": "1234567894", "inspection_date": date(2026, 2, 11),
     "section": "135", "assessment_total": 65000,
     "notice_status": "dispatched", "payment_status": "paid"},
    {"ONLINE NO": "ON2526005", "div no": "UPL02", "Name": "Sunita Yadav",
     "father name": "Ram Yadav", "village": "Tanda",
     "connection no": "1234567895", "inspection_date": date(2026, 3, 22),
     "section": "126", "assessment_total": 240000,
     "notice_status": "pending", "payment_status": "pending"},
    {"ONLINE NO": "ON2526006", "div no": "UPL01", "Name": "गीता देवी",
     "father name": "रामनाथ", "village": "मिलक",
     "connection no": "1234567891", "inspection_date": date(2026, 4, 9),
     "section": "135", "assessment_total": 38500,
     "notice_status": "dispatched", "payment_status": "pending"},
]


# =====================================================================
# 4. DEVICE MASTER  (device list.xlsx)
# =====================================================================
DEVICES = [
    # Lighting
    {"Device Name": "Bulb (Incandescent)",      "Category": "Lighting",   "Load (W)":  60, "Factor": 1.0, "Hours": 6,  "Days": 365, "Unit": "Nos"},
    {"Device Name": "CFL",                      "Category": "Lighting",   "Load (W)":  20, "Factor": 1.0, "Hours": 6,  "Days": 365, "Unit": "Nos"},
    {"Device Name": "LED Bulb",                 "Category": "Lighting",   "Load (W)":   9, "Factor": 1.0, "Hours": 6,  "Days": 365, "Unit": "Nos"},
    {"Device Name": "Tube Light (40W)",         "Category": "Lighting",   "Load (W)":  40, "Factor": 1.0, "Hours": 8,  "Days": 365, "Unit": "Nos"},
    # Cooling
    {"Device Name": "Ceiling Fan",              "Category": "Cooling",    "Load (W)":  75, "Factor": 1.0, "Hours": 12, "Days": 180, "Unit": "Nos"},
    {"Device Name": "Table Fan",                "Category": "Cooling",    "Load (W)":  60, "Factor": 1.0, "Hours": 8,  "Days": 180, "Unit": "Nos"},
    {"Device Name": "Cooler",                   "Category": "Cooling",    "Load (W)": 200, "Factor": 1.0, "Hours": 10, "Days": 120, "Unit": "Nos"},
    {"Device Name": "AC 1.0 Ton",               "Category": "Cooling",    "Load (W)":1200, "Factor": 0.7, "Hours":  8, "Days":  90, "Unit": "Nos"},
    {"Device Name": "AC 1.5 Ton",               "Category": "Cooling",    "Load (W)":1700, "Factor": 0.7, "Hours":  8, "Days":  90, "Unit": "Nos"},
    # Heating
    {"Device Name": "Geyser 15L",               "Category": "Heating",    "Load (W)":2000, "Factor": 0.5, "Hours":  1, "Days": 120, "Unit": "Nos"},
    {"Device Name": "Heater (Room)",            "Category": "Heating",    "Load (W)":2000, "Factor": 0.8, "Hours":  6, "Days":  60, "Unit": "Nos"},
    {"Device Name": "Iron Press",               "Category": "Heating",    "Load (W)":1000, "Factor": 0.5, "Hours":  1, "Days": 365, "Unit": "Nos"},
    # Washing
    {"Device Name": "Washing Machine",          "Category": "Washing",    "Load (W)": 500, "Factor": 0.5, "Hours":  1, "Days": 365, "Unit": "Nos"},
    # Kitchen
    {"Device Name": "Refrigerator (Single Door)","Category": "Kitchen",   "Load (W)": 150, "Factor": 0.4, "Hours": 24, "Days": 365, "Unit": "Nos"},
    {"Device Name": "Refrigerator (Double Door)","Category": "Kitchen",   "Load (W)": 250, "Factor": 0.4, "Hours": 24, "Days": 365, "Unit": "Nos"},
    {"Device Name": "Microwave Oven",           "Category": "Kitchen",    "Load (W)":1200, "Factor": 0.4, "Hours":  1, "Days": 365, "Unit": "Nos"},
    {"Device Name": "Mixer Grinder",            "Category": "Kitchen",    "Load (W)": 500, "Factor": 0.5, "Hours":  1, "Days": 365, "Unit": "Nos"},
    # Pumping
    {"Device Name": "Submersible Pump 1HP",     "Category": "Pumping",    "Load (W)": 750, "Factor": 1.0, "Hours":  6, "Days": 200, "Unit": "Nos"},
    {"Device Name": "Submersible Pump 2HP",     "Category": "Pumping",    "Load (W)":1500, "Factor": 1.0, "Hours":  6, "Days": 200, "Unit": "Nos"},
    {"Device Name": "Submersible Pump 5HP",     "Category": "Pumping",    "Load (W)":3700, "Factor": 1.0, "Hours":  8, "Days": 180, "Unit": "Nos"},
    {"Device Name": "Monoblock Pump 1HP",       "Category": "Pumping",    "Load (W)": 750, "Factor": 1.0, "Hours":  4, "Days": 180, "Unit": "Nos"},
    # Electronics
    {"Device Name": "TV (LED 32\")",            "Category": "Electronics","Load (W)":  60, "Factor": 1.0, "Hours":  6, "Days": 365, "Unit": "Nos"},
    {"Device Name": "TV (LED 43\")",            "Category": "Electronics","Load (W)":  90, "Factor": 1.0, "Hours":  6, "Days": 365, "Unit": "Nos"},
    {"Device Name": "Computer / Desktop",       "Category": "Electronics","Load (W)": 200, "Factor": 0.8, "Hours":  8, "Days": 300, "Unit": "Nos"},
    {"Device Name": "Laptop Charger",           "Category": "Electronics","Load (W)":  65, "Factor": 1.0, "Hours":  6, "Days": 365, "Unit": "Nos"},
    {"Device Name": "Mobile Charger",           "Category": "Electronics","Load (W)":  10, "Factor": 1.0, "Hours":  3, "Days": 365, "Unit": "Nos"},
    # Misc
    {"Device Name": "Inverter Charger",         "Category": "Misc",       "Load (W)": 200, "Factor": 0.5, "Hours":  4, "Days": 200, "Unit": "Nos"},
    {"Device Name": "CCTV (NVR + 4 Cam)",       "Category": "Misc",       "Load (W)":  60, "Factor": 1.0, "Hours": 24, "Days": 365, "Unit": "Nos"},
    {"Device Name": "Water Filter / RO",        "Category": "Misc",       "Load (W)":  60, "Factor": 0.3, "Hours":  1, "Days": 365, "Unit": "Nos"},
    {"Device Name": "Sewing Machine (Electric)","Category": "Misc",       "Load (W)": 100, "Factor": 0.5, "Hours":  4, "Days": 200, "Unit": "Nos"},
]


# =====================================================================
# 5. RATE / SLAB MASTER  (slab_rates.xlsx)
# =====================================================================
RATES = [
    # LMV-1 (Domestic, Rural / Urban)
    {"Category": "LMV-1", "SlabStart":   0, "SlabEnd": 100, "RatePerUnit": 5.50, "FixedCharge": 100, "DutyPercent": 5.0, "Condition": "Domestic 0-100", "EffectiveDate": date(2024, 4, 1)},
    {"Category": "LMV-1", "SlabStart": 101, "SlabEnd": 200, "RatePerUnit": 6.00, "FixedCharge": 100, "DutyPercent": 5.0, "Condition": "Domestic 101-200", "EffectiveDate": date(2024, 4, 1)},
    {"Category": "LMV-1", "SlabStart": 201, "SlabEnd": 300, "RatePerUnit": 6.50, "FixedCharge": 100, "DutyPercent": 5.0, "Condition": "Domestic 201-300", "EffectiveDate": date(2024, 4, 1)},
    {"Category": "LMV-1", "SlabStart": 301, "SlabEnd": None, "RatePerUnit": 7.00, "FixedCharge": 100, "DutyPercent": 5.0, "Condition": "Domestic >300", "EffectiveDate": date(2024, 4, 1)},

    # LMV-2 (Commercial)
    {"Category": "LMV-2", "SlabStart":   0, "SlabEnd": 300, "RatePerUnit": 8.50, "FixedCharge": 250, "DutyPercent": 7.5, "Condition": "Commercial 0-300", "EffectiveDate": date(2024, 4, 1)},
    {"Category": "LMV-2", "SlabStart": 301, "SlabEnd": None, "RatePerUnit": 9.20, "FixedCharge": 250, "DutyPercent": 7.5, "Condition": "Commercial >300", "EffectiveDate": date(2024, 4, 1)},

    # LMV-5 (Agriculture - metered)
    {"Category": "LMV-5", "SlabStart":   0, "SlabEnd": None, "RatePerUnit": 1.75, "FixedCharge":  70, "DutyPercent": 0.0, "Condition": "Pvt Tubewell metered", "EffectiveDate": date(2024, 4, 1)},

    # LMV-6 (Industry small)
    {"Category": "LMV-6", "SlabStart":   0, "SlabEnd": None, "RatePerUnit": 7.30, "FixedCharge": 350, "DutyPercent": 7.5, "Condition": "Small Industry", "EffectiveDate": date(2024, 4, 1)},

    # LMV-9 (Public lamps)
    {"Category": "LMV-9", "SlabStart":   0, "SlabEnd": None, "RatePerUnit": 6.20, "FixedCharge":   0, "DutyPercent": 0.0, "Condition": "Street lights", "EffectiveDate": date(2024, 4, 1)},
]


# =====================================================================
# 6. ACCOUNT MAPPING  (account_mapping.xlsx)
# =====================================================================
MAPPINGS = [
    {"Old Account": "OLD2200001", "New Account": "1234567890", "SC Number": "SC100234",
     "Consumer Name": "रामकुमार सिंह", "Father Name": "श्यामलाल सिंह",
     "Village": "रामपुर", "Effective Date": date(2023, 4, 1), "Status": "active"},
    {"Old Account": "OLD2200002", "New Account": "1234567891", "SC Number": "SC100235",
     "Consumer Name": "गीता देवी", "Father Name": "रामनाथ",
     "Village": "मिलक", "Effective Date": date(2023, 4, 1), "Status": "active"},
    {"Old Account": "OLD2200003", "New Account": "1234567892", "SC Number": "SC100236",
     "Consumer Name": "Mohit Sharma", "Father Name": "Suresh Sharma",
     "Village": "Suar", "Effective Date": date(2023, 4, 1), "Status": "active"},
    {"Old Account": "OLD2200004", "New Account": "1234567893", "SC Number": "SC100237",
     "Consumer Name": "Hari Prakash", "Father Name": "Late Devraj",
     "Village": "Chamraua", "Effective Date": date(2023, 4, 1), "Status": "inactive"},
    {"Old Account": "OLD2200005", "New Account": "1234567894", "SC Number": "SC100238",
     "Consumer Name": "अनिल कुमार", "Father Name": "बद्री प्रसाद",
     "Village": "बिलासपुर", "Effective Date": date(2023, 4, 1), "Status": "active"},
]


# =====================================================================
# Excel writer with styling
# =====================================================================
HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)


def _autosize(ws, df: pd.DataFrame) -> None:
    for i, col in enumerate(df.columns, start=1):
        max_len = max(
            len(str(col)),
            *(len(str(v)) if v is not None else 0 for v in df[col].tolist()),
        )
        ws.column_dimensions[get_column_letter(i)].width = min(max(max_len + 2, 12), 40)


def _style_header(ws) -> None:
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"


def write_workbook(filename: str, sheet_name: str, rows: list[dict],
                   notes: list[str] | None = None) -> Path:
    """Write a single-sheet workbook with header styling and a NOTES sheet."""
    out = MASTER / filename
    df = pd.DataFrame(rows)
    with pd.ExcelWriter(out, engine="openpyxl") as xl:
        df.to_excel(xl, sheet_name=sheet_name, index=False)
        if notes:
            pd.DataFrame({"NOTES": notes}).to_excel(xl, sheet_name="README", index=False)
    # Apply styling
    wb = load_workbook(out)
    ws = wb[sheet_name]
    _style_header(ws)
    _autosize(ws, df)
    if "README" in wb.sheetnames:
        rws = wb["README"]
        _style_header(rws)
        rws.column_dimensions["A"].width = 110
        for r in range(2, rws.max_row + 1):
            rws.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(out)
    return out


def main() -> None:
    files = [
        ("SAMPLE_raid_master_data.xlsx", "Consumers", CONSUMERS, [
            "TARGET TABLE: consumers (69k rows in production)",
            "REQUIRED COLUMN: ACCT_ID (any of: account_id, k_no, service_no, connection_no, consumer_id)",
            "Optional: NAME, FATHER_NAME, ADDRESS, VILLAGE, MOBILE_NO, LOAD, LOAD_UNIT, SUPPLY_TYPE, CATEGORY, CON_STATUS, DIV_CODE, SC_NO, etc.",
            "Importer normalises column names (case/space/punct insensitive) so 'Account ID', 'ACCT_ID', 'A/C No' all map to account_number.",
            "Hindi (Devanagari) text is fully supported; Krutidev font text will be stored verbatim — convert to Unicode for best search.",
        ]),
        ("SAMPLE_all_data.xlsx", "Historical", HISTORICAL, [
            "TARGET TABLE: historical_cases (8,956 rows in production)",
            "Columns matched: div no, Name, father name, village, Account Id, Date, assessment, FIR, dhara",
            "No required column — but rows with neither name nor account_id will be skipped.",
            "Date can be Excel date, dd/mm/yyyy or yyyy-mm-dd — auto-parsed.",
        ]),
        ("SAMPLE_raid_excell_2526.xlsx", "Current", CURRENT, [
            "TARGET TABLE: current_cases (24,195 rows in production)",
            "Recommended: ONLINE NO is unique key — same online_no will UPDATE existing row.",
            "Columns: ONLINE NO, div no, Name, father name, village, connection no, inspection_date, section, assessment_total, notice_status, payment_status",
            "section values: 135 (theft) / 138 (TD) / 126 (UUE) / Other",
        ]),
        ("SAMPLE_device_list.xlsx", "Devices", DEVICES, [
            "TARGET TABLE: device_master (38 default in spec, 40 auto-seeded on first run)",
            "REQUIRED: Device Name (unique). Existing rows are UPDATED in place.",
            "Optional: Category, Load (W), Factor, Hours, Days, Unit",
            "Default Factor=1.0, Hours=8, Days=365 if omitted.",
            "Categories: Lighting, Cooling, Heating, Washing, Kitchen, Pumping, Electronics, Misc",
        ]),
        ("SAMPLE_slab_rates.xlsx", "Rates", RATES, [
            "TARGET TABLE: rate_master (24 rows in production)",
            "REQUIRED: Category (LMV-1 ... LMV-9)",
            "SlabEnd blank = unlimited (open-ended top slab).",
            "DutyPercent is electricity-duty %: 5 / 7.5 / 0 typical.",
            "Multiple rows per category build the slab ladder (e.g. 0-100, 101-200, 201+).",
            "EffectiveDate lets you keep tariff history.",
        ]),
        ("SAMPLE_account_mapping.xlsx", "Mapping", MAPPINGS, [
            "TARGET TABLE: account_mapping (optional — bridges old↔new account numbers)",
            "Used by 4-level offense detection (Level 3 lookup).",
            "Columns: Old Account, New Account, SC Number, Consumer Name, Father Name, Village, Effective Date, Status",
            "At least one of Old Account / New Account must be present.",
        ]),
    ]
    for fname, sheet, rows, notes in files:
        path = write_workbook(fname, sheet, rows, notes)
        print(f"  wrote {path.relative_to(ROOT)}  ({len(rows)} rows)")
    print(f"\nAll {len(files)} sample workbooks written to {MASTER.relative_to(ROOT)}/")


if __name__ == "__main__":
    main()
